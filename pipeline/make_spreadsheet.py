#!/usr/bin/env python3
"""
Build a normalized spreadsheet from extracted project JSONs.

Usage:
    python make_spreadsheet.py
    python make_spreadsheet.py --extracted_dir extracted/ --output projects.xlsx

Produces two sheets:
  Projects  — one row per project, scalar fields
  Team      — one row per (project, role, firm), joinable to Projects on project_id
"""

import argparse
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Maps JSON field names → display labels for both teams
DESIGN_ROLE_LABELS = {
    "architect": "Architect",
    "structural_engineer": "Structural Engineer",
    "mechanical_engineer": "Mechanical Engineer",
    "electrical_engineer": "Electrical Engineer",
    "civil_engineer": "Civil Engineer",
    "interior_design": "Interior Design",
    "landscape_architect": "Landscape Architect",
    "geotech_engineer": "Geotech Engineer",
    "lighting_design": "Lighting Design",
    "food_service_design": "Food Service Design",
    "furniture": "Furniture",
}

CONSTRUCTION_ROLE_LABELS = {
    "general_contractor": "General Contractor",
    "plumbing": "Plumbing",
    "hvac": "HVAC",
    "electrical": "Electrical",
    "concrete": "Concrete",
    "steel_fabrication": "Steel Fabrication",
    "steel_erection": "Steel Erection",
    "glass_curtain_wall": "Glass / Curtain Wall",
    "masonry": "Masonry",
    "drywall_acoustics": "Drywall / Acoustics",
    "painting": "Painting",
    "tile_stone": "Tile / Stone",
    "carpentry": "Carpentry",
    "flooring": "Flooring",
    "roofing": "Roofing",
    "waterproofing": "Waterproofing",
    "excavation": "Excavation",
    "demolition": "Demolition",
    "landscaping": "Landscaping",
    "millwork": "Millwork",
}


def parse_role_firm(entry: str) -> tuple[str, str]:
    """Split 'Role: Firm Name' into (role, firm). Falls back to ('Other', entry)."""
    if ":" in entry:
        role, _, firm = entry.partition(":")
        return role.strip(), firm.strip()
    return "Other", entry.strip()


def iter_team_rows(project_id: int, project: dict) -> list[dict]:
    rows = []

    def add(team: str, role: str, firms_raw: str | None):
        if not firms_raw:
            return
        for firm in re.split(r"\s*/\s*", firms_raw):
            firm = firm.strip()
            if firm:
                rows.append({"project_id": project_id, "team": team, "role": role, "firm": firm})

    design = project.get("design_team") or {}
    for field, label in DESIGN_ROLE_LABELS.items():
        add("Design", label, design.get(field))
    for entry in design.get("other") or []:
        role, firm = parse_role_firm(entry)
        add("Design", role, firm)

    construction = project.get("construction_team") or {}
    for field, label in CONSTRUCTION_ROLE_LABELS.items():
        add("Construction", label, construction.get(field))
    for entry in construction.get("other") or []:
        role, firm = parse_role_firm(entry)
        add("Construction", role, firm)

    return rows


def load_projects(extracted_dir: Path) -> list[dict]:
    projects = []
    for f in sorted(extracted_dir.glob("*.json")):
        for p in json.loads(f.read_text()):
            projects.append(p)
    return projects


def style_header_row(ws, row: int, fill_hex: str):
    fill = PatternFill("solid", fgColor=fill_hex)
    bold = Font(bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = bold
        cell.alignment = Alignment(wrap_text=False)


def autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


def build_workbook(projects: list[dict]) -> Workbook:
    wb = Workbook()

    # ── Projects sheet ──────────────────────────────────────────────
    ws_proj = wb.active
    ws_proj.title = "Projects"

    proj_headers = [
        "project_id", "project_name", "location", "cost", "delivery_method",
        "stories_levels", "square_footage", "year_completed",
        "owner", "owner_rep", "developer",
        "source_file", "source_page",
    ]
    ws_proj.append(proj_headers)
    style_header_row(ws_proj, 1, "1F4E79")
    for cell in ws_proj[1]:
        cell.font = Font(bold=True, color="FFFFFF")

    for pid, p in enumerate(projects, start=1):
        ws_proj.append([
            pid,
            p.get("project_name"),
            p.get("location"),
            p.get("cost"),
            p.get("delivery_method"),
            p.get("stories_levels"),
            p.get("square_footage"),
            p.get("year_completed"),
            p.get("owner"),
            p.get("owner_rep"),
            p.get("developer"),
            p.get("source_file"),
            p.get("source_page"),
        ])

    autofit(ws_proj)

    # ── Team sheet ───────────────────────────────────────────────────
    ws_team = wb.create_sheet("Team")

    team_headers = ["project_id", "project_name", "team", "role", "firm"]
    ws_team.append(team_headers)
    style_header_row(ws_team, 1, "375623")
    for cell in ws_team[1]:
        cell.font = Font(bold=True, color="FFFFFF")

    for pid, p in enumerate(projects, start=1):
        for row in iter_team_rows(pid, p):
            ws_team.append([
                row["project_id"],
                p.get("project_name"),
                row["team"],
                row["role"],
                row["firm"],
            ])

    autofit(ws_team)

    return wb


def main():
    parser = argparse.ArgumentParser(description="Build spreadsheet from extracted project JSONs")
    parser.add_argument("--extracted_dir", default="extracted", help="Directory of per-issue JSONs (default: extracted/)")
    parser.add_argument("--output", "-o", default="projects.xlsx", help="Output Excel file (default: projects.xlsx)")
    args = parser.parse_args()

    extracted_dir = Path(args.extracted_dir)
    if not extracted_dir.exists():
        print(f"ERROR: {extracted_dir} not found")
        return

    projects = load_projects(extracted_dir)
    print(f"Loaded {len(projects)} project(s) from {extracted_dir}")

    wb = build_workbook(projects)
    wb.save(args.output)
    print(f"Saved → {args.output}")


if __name__ == "__main__":
    main()
